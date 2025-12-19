import sys
sys.path.append(r'c:\Users\Propietario\Downloads\makinas')
from maquinas_app import detect_header_row
import pandas as pd
from openpyxl import load_workbook

path = r'c:\Users\Propietario\Downloads\makinas\Matriz de condición de equipos principales excel.xlsx'
header = detect_header_row(path, 'CM Matrix equipos principales')
print('Detected header row (0-based):', header)
df = pd.read_excel(path, sheet_name='CM Matrix equipos principales', header=header)
print('Columns:')
for i,c in enumerate(df.columns):
    print(i, repr(str(c)))

wb = load_workbook(path, data_only=True)
ws = wb['CM Matrix equipos principales']
# show first 5 data rows cell fills for columns 0..15
excel_header_row = header+1
start = excel_header_row+1
print('\nCell fills for first data rows (hex if available):')
for r in range(start, start+5):
    print('Row', r)
    for c in range(1, 16):
        cell = ws.cell(row=r, column=c)
        fg = None
        if cell.fill and hasattr(cell.fill, 'fgColor'):
            fg = cell.fill.fgColor.rgb or getattr(cell.fill.start_color, 'rgb', None) or getattr(cell.fill.start_color, 'index', None)
        if fg:
            print(c, fg, end=' | ')
    print('\n')
print('Done')
