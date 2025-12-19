"""
Monitor de Condición - Seguimiento de máquinas con herramientas de diagnóstico
- SQLite (machines.db)
- CRUD máquinas, herramientas, mediciones
- Resumen por máquina
"""

from flask import Flask, request, redirect, url_for, render_template_string, flash
from werkzeug.security import generate_password_hash
import sqlite3, os
from datetime import datetime
import pandas as pd
import math
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = "secret_key"

DB_FILE = "machines.db"

# ---- Database ----
def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    
    # Create tables if they don't exist
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS machines (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        notes TEXT,
        priority INTEGER DEFAULT 3,
        machine_group INTEGER DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS tools (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        description TEXT
    );
    CREATE TABLE IF NOT EXISTS measurements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        machine_id INTEGER,
        tool_id INTEGER,
        date TEXT,
        criticality INTEGER,
        note TEXT,
        FOREIGN KEY(machine_id) REFERENCES machines(id),
        FOREIGN KEY(tool_id) REFERENCES tools(id)
    );
    """)
    
    # Migration: Add missing columns to existing machines table
    cursor = conn.execute("PRAGMA table_info(machines)")
    machines_columns = {row[1] for row in cursor.fetchall()}
    
    if 'priority' not in machines_columns:
        conn.execute("ALTER TABLE machines ADD COLUMN priority INTEGER DEFAULT 3")
        print("✓ Columna priority agregada")
    
    if 'machine_group' not in machines_columns:
        conn.execute("ALTER TABLE machines ADD COLUMN machine_group INTEGER DEFAULT 1")
        print("✓ Columna machine_group agregada")
    
    if 'notes' not in machines_columns:
        conn.execute("ALTER TABLE machines ADD COLUMN notes TEXT")
        print("✓ Columna notes agregada")
    if 'hac_code' not in machines_columns:
      conn.execute("ALTER TABLE machines ADD COLUMN hac_code TEXT")
      print("✓ Columna hac_code agregada")
      machines_columns.add('hac_code')
    
    # Migration: Add missing columns to existing measurements table
    cursor = conn.execute("PRAGMA table_info(measurements)")
    measurements_columns = {row[1] for row in cursor.fetchall()}
    
    if 'criticality' not in measurements_columns:
        conn.execute("ALTER TABLE measurements ADD COLUMN criticality INTEGER")
        print("✓ Columna criticality agregada")
    
    if 'note' not in measurements_columns:
        conn.execute("ALTER TABLE measurements ADD COLUMN note TEXT")
        print("✓ Columna note agregada")
    if 'severity' not in measurements_columns:
      conn.execute("ALTER TABLE measurements ADD COLUMN severity TEXT")
      print("✓ Columna severity agregada")
    if 'repair_time' not in measurements_columns:
      conn.execute("ALTER TABLE measurements ADD COLUMN repair_time TEXT")
      print("✓ Columna repair_time agregada")
    
    # Migration: Add missing columns to existing tools table
    cursor = conn.execute("PRAGMA table_info(tools)")
    tools_columns = {row[1] for row in cursor.fetchall()}
    
    if 'description' not in tools_columns:
        conn.execute("ALTER TABLE tools ADD COLUMN description TEXT")
        print("✓ Columna description agregada")
    # Migration: add color and color_hex columns to machines
    if 'color' not in machines_columns:
      conn.execute("ALTER TABLE machines ADD COLUMN color TEXT")
      print("✓ Columna color agregada")
      machines_columns.add('color')
    if 'color_hex' not in machines_columns:
      conn.execute("ALTER TABLE machines ADD COLUMN color_hex TEXT")
      print("✓ Columna color_hex agregada")
      machines_columns.add('color_hex')
    if 'machine_type' not in machines_columns:
      conn.execute("ALTER TABLE machines ADD COLUMN machine_type TEXT")
      print("✓ Columna machine_type agregada")
      machines_columns.add('machine_type')
    
    conn.commit()
    conn.close()
    print("✓ Base de datos inicializada")

init_db()

# ---- Base Template ----
BASE = """
<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{{ page_title }}</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
<style>
.root{ }
:root{--bg:#f5f7fb;--accent-1:#2563eb;--accent-2:#1e40af;--card-bg:#ffffff;--text:#0f172a;--muted:#6b7280}
html.dark{--bg:#0b1220;--card-bg:#071028;--accent-1:#0ea5e9;--accent-2:#0369a1;--text:#e6eef8;--muted:#94a3b8}
body { background:var(--bg); color:var(--text); font-family: Inter, system-ui, -apple-system, 'Segoe UI', Roboto, 'Helvetica Neue', Arial; }
.navbar { background: linear-gradient(90deg, var(--accent-1), var(--accent-2)); }
.card { border-radius: 12px; box-shadow: 0 6px 18px rgba(16,24,40,0.06); background: var(--card-bg); }
.table thead th { background: var(--accent-1); color: white; }
.crit-1 { background: #d1fae5; } .crit-1-text { color: #047857; }
.crit-2 { background: #dbeafe; } .crit-2-text { color: #1e40af; }
.crit-3 { background: #fef3c7; } .crit-3-text { color: #b45309; }
.crit-4 { background: #fee2e2; } .crit-4-text { color: #991b1b; }
.badge-crit { padding: 6px 10px; border-radius: 6px; font-weight: 600; }
/* Severity badges */
.sev-rojo{background:#fee2e2;color:#991b1b;padding:6px 8px;border-radius:6px}
.sev-naranja{background:#fff4e6;color:#9a5800;padding:6px 8px;border-radius:6px}
.sev-amarillo{background:#fff7cc;color:#7a4b00;padding:6px 8px;border-radius:6px}
.sev-verde{background:#e6ffef;color:#065f46;padding:6px 8px;border-radius:6px}
.sev-gris{background:#f3f4f6;color:#374151;padding:6px 8px;border-radius:6px}
.card-hover:hover{transform:translateY(-4px);transition:all .18s ease}
.machine-card { border-left: 6px solid transparent; }
.small-muted{color:var(--muted);font-size:0.85rem}
</style>
</head>
<body>
<nav class="navbar navbar-expand-lg navbar-dark">
    <div class="container-fluid">
    <a class="navbar-brand fw-bold" href="/">📊 Monitor</a>
    <div class="navbar-nav ms-auto">
      <a class="nav-link" href="/">Máquinas</a>
      <a class="nav-link" href="/tools">Herramientas</a>
      <a class="nav-link" href="/calendar">Calendario</a>
      <button id="theme_toggle" class="btn btn-sm btn-outline-light ms-2" title="Alternar modo" type="button">🌙</button>
    </div>
  </div>
</nav>

<div class="container mt-4">
{{ body }}
</div>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
"""

def render(body_template, **context):
    page_title = context.pop('page_title', 'Monitor')
    full_template = BASE.replace('{{ body }}', body_template)
    return render_template_string(full_template, page_title=page_title, **context)

# ============ MÁQUINAS ============

MACHINES_LIST = """
<div class="d-flex justify-content-between align-items-center mb-3">
  <h3>Máquinas</h3>
  <a class="btn btn-primary" href="/machines/add">+ Agregar</a>
</div>
<div class="mb-2"><small class="text-muted">Agrupado por: Tipo de equipo (según Excel)</small></div>

<div class="card p-4 mb-4" style="background: #f8f9fa;">
  <h6 class="mb-3"><strong>Filtrar Máquinas</strong></h6>
  <form method="get" class="row g-3">
    <div class="col-md-4">
      <label class="form-label">Buscar por nombre</label>
      <input type="text" class="form-control" name="search" placeholder="Nombre de máquina..." value="{{ search or '' }}">
    </div>
    <div class="col-md-3">
      <label class="form-label">Prioridad</label>
      <select class="form-select" name="priority">
        <option value="">-- Todas --</option>
        <option value="1" {% if filter_priority == '1' %}selected{% endif %}>1 (Baja)</option>
        <option value="2" {% if filter_priority == '2' %}selected{% endif %}>2</option>
        <option value="3" {% if filter_priority == '3' %}selected{% endif %}>3 (Media)</option>
        <option value="4" {% if filter_priority == '4' %}selected{% endif %}>4</option>
        <option value="5" {% if filter_priority == '5' %}selected{% endif %}>5 (Alta)</option>
      </select>
    </div>
      
    <div class="col-md-2 d-flex align-items-end">
      <button type="submit" class="btn btn-primary w-100">Filtrar</button>
    </div>
  </form>
  {% if search or filter_priority or filter_group %}
  <div class="mt-2">
    <a href="/" class="btn btn-sm btn-outline-secondary">Limpiar filtros</a>
  </div>
  {% endif %}
</div>

<style>
  .group-header {
    cursor: pointer;
    user-select: none;
    padding: 12px;
    background: #e9ecef;
    border-radius: 6px;
    margin-top: 20px;
    display: flex;
    align-items: center;
    gap: 10px;
  }
  .group-header:hover {
    background: #dee2e6;
  }
  .chevron {
    transition: transform 0.2s;
    font-size: 18px;
  }
  .chevron.collapsed {
    transform: rotate(-90deg);
  }
  .group-content {
    display: none;
  }
  .group-content.show {
    display: block;
  }
</style>

{% for type_name, group_machines in groups|dictsort %}
  <div class="group-header" onclick="toggleGroup(this, 'group-{{ loop.index }}')">
    <span class="chevron collapsed">▶</span>
    <strong>{{ type_name }}</strong> <span class="text-muted">({{ group_machines | length }})</span>
  </div>

  <div id="group-{{ loop.index }}" class="group-content">
    <div class="row mt-3">
      {% for m in group_machines %}
      <div class="col-md-4 mb-3">
        <div class="card p-3 card-hover machine-card" {% if m.color_hex %} style="border-left:6px solid #{{ m.color_hex }}" {% endif %}>
          <h5><a href="/machines/{{ m.id }}" class="text-decoration-none">{{ m.name }}</a></h5>
          {% if m.hac_code %}<small class="text-muted">HAC: {{ m.hac_code }}</small>{% endif %}
          <div class="mb-2 d-flex align-items-start gap-2">
            <span class="badge bg-warning">Prioridad: {{ m.priority }}/5</span>
            {% if m.color_hex %}
              <span class="badge" style="display:inline-block;width:18px;height:18px;border-radius:4px;background:#{{ m.color_hex }};border:1px solid #ccc;margin-left:8px;vertical-align:middle;"></span>
              <small class="ms-2 small-muted">{{ {'red':'Rojo','yellow':'Amarillo','blue':'Azul','green':'Verde'}.get(m.color, m.color or 'Sin color') }}</small>
              {% if m.color == 'red' %}
                <div class="text-danger small">Se tiene que arreglar</div>
              {% elif m.color == 'yellow' %}
                <div class="text-warning small">Revisar</div>
              {% elif m.color == 'blue' %}
                <div class="text-primary small">Atención</div>
              {% elif m.color == 'green' %}
                <div class="text-success small">Bien</div>
              {% endif %}
            {% else %}
              <span class="badge bg-secondary ms-2">{{ {'red':'Rojo','yellow':'Amarillo','blue':'Azul','green':'Verde'}.get(m.color, m.color or 'Sin color') }}</span>
              {% if m.color == 'red' %}
                <div class="text-danger small">Se tiene que arreglar</div>
              {% elif m.color == 'yellow' %}
                <div class="text-warning small">Revisar</div>
              {% elif m.color == 'blue' %}
                <div class="text-primary small">Atención</div>
              {% elif m.color == 'green' %}
                <div class="text-success small">Bien</div>
              {% endif %}
            {% endif %}
          </div>
          {% if m.notes %}<small class="text-muted d-block mb-2">{{ m.notes }}</small>{% endif %}
          <div class="mt-2">
            <a class="btn btn-sm btn-outline-primary" href="/machines/{{ m.id }}/edit">Editar</a>
            <a class="btn btn-sm btn-outline-danger" href="/machines/{{ m.id }}/delete" onclick="return confirm('¿Eliminar?')">Eliminar</a>
          </div>
        </div>
      </div>
      {% endfor %}
    </div>
  </div>
{% endfor %}

{% if not machines %}
<div class="alert alert-info mt-4">No se encontraron máquinas con los filtros aplicados</div>
{% endif %}

<script>
function toggleGroup(header, contentId) {
  const content = document.getElementById(contentId);
  const chevron = header.querySelector('.chevron');
  
  content.classList.toggle('show');
  chevron.classList.toggle('collapsed');
}
</script>
"""

MACHINE_ADD = """
<div class="row justify-content-center">
  <div class="col-md-6">
    <h4>Agregar Máquina</h4>
    <form method="post" class="card p-4">
      <label class="form-label"><strong>Nombre</strong></label>
      <input class="form-control mb-3" name="name" placeholder="Nombre de la máquina" required>
      
      <label class="form-label"><strong>Grupo (1-5)</strong></label>
      <select name="machine_group" class="form-select mb-3" required>
        <option value="1">Grupo 1</option>
        <option value="2">Grupo 2</option>
        <option value="3">Grupo 3</option>
        <option value="4">Grupo 4</option>
        <option value="5">Grupo 5</option>
      </select>
      
      <label class="form-label"><strong>Prioridad (1-5)</strong></label>
      <input type="number" name="priority" min="1" max="5" value="3" class="form-control mb-3" required>
      
      <label class="form-label"><strong>Notas</strong></label>
      <textarea name="notes" class="form-control mb-3" rows="3" placeholder="Notas adicionales..."></textarea>
      
      <button class="btn btn-primary w-100">Guardar</button>
    </form>
  </div>
</div>
"""

MACHINE_EDIT = """
<div class="row justify-content-center">
  <div class="col-md-6">
    <h4>Editar Máquina</h4>
    <form method="post" class="card p-4">
      <label class="form-label"><strong>Nombre</strong></label>
      <input class="form-control mb-3" name="name" value="{{ m.name }}" required>
      
      <label class="form-label"><strong>Prioridad (1-5)</strong></label>
      <input type="number" name="priority" min="1" max="5" value="{{ m.priority }}" class="form-control mb-3" required>
      
      <label class="form-label"><strong>Notas</strong></label>
      <textarea name="notes" class="form-control mb-3" rows="3">{{ m.notes or '' }}</textarea>
      
      <button class="btn btn-primary w-100">Guardar</button>
    </form>
  </div>
</div>
"""

MACHINE_DETAIL = """
<div class="d-flex justify-content-between align-items-center mb-3">
  <div>
    <h3>{{ machine.name }}</h3>
    {% if machine.hac_code %}<small class="text-muted">HAC: {{ machine.hac_code }}</small>{% endif %}
    <div class="mt-2">
      <span class="badge bg-warning">Prioridad: {{ machine.priority }}/5</span>
      {% if machine['color_hex'] %}
        <span class="badge" style="display:inline-block;width:16px;height:16px;border-radius:4px;background:#{{ machine['color_hex'] }};border:1px solid #ccc;margin-left:8px;vertical-align:middle;"></span>
        <small class="ms-2">{{ {'red':'Rojo','yellow':'Amarillo','blue':'Azul','green':'Verde'}.get(machine['color'], machine['color'] or 'Sin color') }}</small>
      {% else %}
        {% if machine['color'] %}
          <span class="badge bg-secondary ms-2">{{ {'red':'Rojo','yellow':'Amarillo','blue':'Azul','green':'Verde'}.get(machine['color'], machine['color']) }}</span>
        {% endif %}
      {% endif %}
      {% if machine['color'] == 'red' %}
        <div class="text-danger small">Se tiene que arreglar</div>
      {% elif machine['color'] == 'yellow' %}
        <div class="text-warning small">Revisar</div>
      {% elif machine['color'] == 'blue' %}
        <div class="text-primary small">Atención</div>
      {% elif machine['color'] == 'green' %}
        <div class="text-success small">Bien</div>
      {% endif %}
      {% if machine.notes %}<small class="text-muted d-block mt-2">{{ machine.notes }}</small>{% endif %}
    </div>
  </div>
  <div>
    <a class="btn btn-primary" href="/measurements/add?mid={{ machine.id }}">+ Medición</a>
    <a class="btn btn-outline-secondary" href="/machines/{{ machine.id }}/edit">Editar</a>
    <a class="btn btn-secondary" href="/">Atrás</a>
  </div>
</div>

<h5 class="mt-4 mb-3">Estado Actual (última medición por herramienta)</h5>
<div class="row">
  {% for tool_status in current_status %}
  <div class="col-md-4 mb-3">
    <div class="card p-3">
      <h6>{{ tool_status.tool }}</h6>
      {% if tool_status.criticality %}
        <div class="badge-crit crit-{{ [tool_status.criticality//3, 1]|max|min(4) }} crit-{{ [tool_status.criticality//3, 1]|max|min(4) }}-text">
          Criticidad: {{ tool_status.criticality }}/10
        </div>
        <small class="text-muted d-block mt-2">{{ tool_status.date }}</small>
        {% if tool_status.note %}<small class="d-block">{{ tool_status.note }}</small>{% endif %}
      {% else %}
        <small class="text-muted">Sin mediciones</small>
      {% endif %}
    </div>
  </div>
  {% endfor %}
</div>

<h5 class="mt-5 mb-3">Historial Completo</h5>
<div class="table-responsive">
  <table class="table">
    <thead>
      <tr><th>Fecha</th><th>Herramienta</th><th>Criticidad</th><th>Nota</th><th></th></tr>
    </thead>
    <tbody>
      {% for record in history %}
      <tr>
        <td>{{ record.date }}</td>
        <td>{{ record.tool }}</td>
        <td><span class="badge crit-{{ [record.criticality//3, 1]|max|min(4) }}">{{ record.criticality }}</span></td>
        <td>{{ record.note or '' }}</td>
        <td><a href="/measurements/{{ record.id }}/delete" class="btn btn-sm btn-outline-danger" onclick="return confirm('¿Eliminar?')">X</a></td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
</div>
"""

CALENDAR_TEMPLATE = """
<div class="d-flex justify-content-between align-items-center mb-3">
  <h3>Calendario - Añadir notas</h3>
  <a class="btn btn-secondary" href="/">Atrás</a>
</div>

<div class="card p-4 mb-4">
  <form method="post" class="row g-3">
    <div class="col-md-3">
      <label class="form-label">Fecha</label>
      <input type="date" name="date" class="form-control" value="{{ today }}" required>
    </div>
    <div class="col-md-4">
      <label class="form-label">Herramientas (selecciona una o varias)</label>
      <select name="tool_id" class="form-select" multiple size="4">
        {% for t in tools %}
        <option value="{{ t.id }}">{{ t.name }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="col-md-2">
      <label class="form-label">Severidad</label>
      <select name="severity" class="form-select">
        <option value="gris">Gris (No aplica)</option>
        <option value="verde">Verde</option>
        <option value="amarillo">Amarillo</option>
        <option value="naranja">Naranja</option>
        <option value="rojo">Rojo</option>
      </select>
      <small class="text-muted">El tiempo de arreglo se asigna automáticamente</small>
    </div>
    <div class="col-md-5">
      <label class="form-label">Nota</label>
      <textarea name="note" class="form-control" rows="3" placeholder="Texto de la nota..."></textarea>
    </div>
    <div class="col-12">
      <label class="form-label">Máquinas (selecciona una, varias o usar "Seleccionar todas")</label>
      <div class="mb-2"><button type="button" class="btn btn-sm btn-outline-secondary" id="select_all_btn">Seleccionar todas</button> <button type="button" class="btn btn-sm btn-outline-secondary" id="clear_all_btn">Limpiar</button></div>
      <div class="row" style="max-height:300px; overflow:auto; border:1px solid #eee; padding:10px; border-radius:6px;">
        {% for m in machines %}
        <div class="col-md-4">
          <label class="form-check">
            <input class="form-check-input machine-checkbox" type="checkbox" name="machine_id" value="{{ m.id }}">
            <span class="form-check-label">{{ m.name }} {% if m.machine_type %} <small class="text-muted">({{ m.machine_type }})</small>{% endif %}{% if m.hac_code %} <small class="text-muted">- HAC: {{ m.hac_code }}</small>{% endif %}</span>
          </label>
        </div>
        {% endfor %}
      </div>
    </div>
    <div class="col-12 text-end">
      <button class="btn btn-primary">Añadir nota</button>
    </div>
  </form>
</div>

<h5>Notas recientes</h5>
<div class="table-responsive card p-3">
  <table class="table">
    <thead><tr><th>Fecha</th><th>Máquina</th><th>HAC</th><th>Herramienta</th><th>Severidad</th><th>Tiempo arreglo</th><th>Nota</th></tr></thead>
    <tbody>
      {% for n in recent %}
      <tr>
        <td class="small-muted">{{ n.date }}</td>
        <td><strong>{{ n.machine }}</strong></td>
        <td class="small-muted">{{ n.hac or '' }}</td>
        <td>{{ n.tool }}</td>
        <td>
          {% if n.severity == 'rojo' %}<span class="sev-rojo">Rojo</span>
          {% elif n.severity == 'naranja' %}<span class="sev-naranja">Naranja</span>
          {% elif n.severity == 'amarillo' %}<span class="sev-amarillo">Amarillo</span>
          {% elif n.severity == 'verde' %}<span class="sev-verde">Verde</span>
          {% else %}<span class="sev-gris">Gris</span>{% endif %}
        </td>
        <td class="small-muted">{{ n.repair_time or '' }}</td>
        <td>{{ n.note or '' }}</td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
</div>

<script>
document.getElementById('select_all_btn').addEventListener('click', function(){
  document.querySelectorAll('.machine-checkbox').forEach(cb => cb.checked = true);
});
document.getElementById('clear_all_btn').addEventListener('click', function(){
  document.querySelectorAll('.machine-checkbox').forEach(cb => cb.checked = false);
});
</script>
"""

@app.route("/")
def machines_list():
  conn = get_db()

  # Get filter parameters
  search = request.args.get('search', '').strip().lower()
  filter_priority = request.args.get('priority', '')
  filter_group = request.args.get('group', '')
  filter_color = request.args.get('color', '')

  # Build SQL query with filters
  query = "SELECT * FROM machines WHERE 1=1"
  params = []

  if search:
    query += " AND LOWER(name) LIKE ?"
    params.append(f"%{search}%")

  if filter_priority:
    query += " AND priority = ?"
    params.append(int(filter_priority))

  if filter_group:
    query += " AND machine_group = ?"
    params.append(int(filter_group))

  query += " ORDER BY priority DESC, name"

  rows = conn.execute(query, params).fetchall()
  # Build machines list with latest criticality and color
  machines = []
  for r in rows:
    mid = r['id']
    latest = conn.execute("SELECT criticality FROM measurements WHERE machine_id=? ORDER BY date DESC LIMIT 1", (mid,)).fetchone()
    latest_crit = latest['criticality'] if latest and latest['criticality'] is not None else None
    # determine color: prefer persisted DB color if present
    if 'color' in r.keys() and r['color']:
      color = r['color']
    else:
      if latest_crit is None:
        color = 'green'
      else:
        try:
          lc = int(latest_crit)
        except Exception:
          lc = 0
        if lc >= 8:
          color = 'red'
        elif lc >= 5:
          color = 'yellow'
        elif lc >= 3:
          color = 'blue'
        else:
          color = 'green'
    machines.append({
      'id': r['id'],
      'name': r['name'],
      'notes': r['notes'],
      'priority': r['priority'],
      'machine_group': (r['machine_group'] if 'machine_group' in r.keys() else 1),
      'latest_crit': latest_crit,
      'color': color,
      'machine_type': (r['machine_type'] if 'machine_type' in r.keys() and r['machine_type'] else 'Sin tipo'),
      'color_hex': (r['color_hex'] if 'color_hex' in r.keys() else None),
      'hac_code': (r['hac_code'] if 'hac_code' in r.keys() else None)
    })

  conn.close()

  # group machines by Tipo equipo for template
  groups = {}
  for m in machines:
    groups.setdefault(m.get('machine_type','Sin tipo'), []).append(m)

  filter_priority_val = filter_priority if filter_priority else None
  filter_group_val = int(filter_group) if filter_group else None

  return render(MACHINES_LIST, page_title="Máquinas", groups=groups, machines=machines,
          search=search, filter_priority=filter_priority_val, filter_group=filter_group_val, filter_color=filter_color)

@app.route("/machines/add", methods=["GET","POST"])
def machines_add():
    if request.method == "POST":
        name = request.form.get("name","").strip()
        priority = int(request.form.get("priority", "3"))
        notes = request.form.get("notes","").strip()
        machine_group = int(request.form.get("machine_group", "1"))
        conn = get_db()
        try:
            conn.execute("INSERT INTO machines (name, priority, notes, machine_group) VALUES (?,?,?,?)", (name, priority, notes, machine_group))
            conn.commit()
            conn.close()
            return redirect("/")
        except:
            conn.close()
            flash("Máquina duplicada o error")
            return render(MACHINE_ADD, page_title="Agregar Máquina")
    return render(MACHINE_ADD, page_title="Agregar Máquina")

@app.route("/machines/<int:id>/edit", methods=["GET","POST"])
def machines_edit(id):
    conn = get_db()
    m = conn.execute("SELECT * FROM machines WHERE id=?", (id,)).fetchone()
    if not m:
        conn.close()
        return "No encontrada", 404
    if request.method == "POST":
        name = request.form.get("name","").strip()
        priority = int(request.form.get("priority", "3"))
        notes = request.form.get("notes","").strip()
        conn.execute("UPDATE machines SET name=?, priority=?, notes=? WHERE id=?", (name, priority, notes, id))
        conn.commit()
        conn.close()
        return redirect(f"/machines/{id}")
    conn.close()
    return render(MACHINE_EDIT, page_title="Editar", m=m)

@app.route("/machines/<int:id>/delete")
def machines_delete(id):
    conn = get_db()
    conn.execute("DELETE FROM measurements WHERE machine_id=?", (id,))
    conn.execute("DELETE FROM machines WHERE id=?", (id,))
    conn.commit()
    conn.close()
    return redirect("/")

@app.route("/machines/<int:id>/move")
def machines_move(id):
    group = request.args.get('group')
    if not group:
        return redirect("/")
    
    try:
        group = int(group)
        if group < 1 or group > 5:
            return redirect("/")
    except:
        return redirect("/")
    
    conn = get_db()
    conn.execute("UPDATE machines SET machine_group=? WHERE id=?", (group, id))
    conn.commit()
    conn.close()
    return redirect("/")

@app.route("/machines/<int:id>")
def machine_detail(id):
    conn = get_db()
    machine = conn.execute("SELECT * FROM machines WHERE id=?", (id,)).fetchone()
    if not machine:
        conn.close()
        return "No encontrada", 404
    
    # Última medición por herramienta
    current = conn.execute("""
        SELECT DISTINCT
            t.id, t.name as tool,
            (SELECT criticality FROM measurements WHERE tool_id=t.id AND machine_id=? ORDER BY date DESC LIMIT 1) as criticality,
            (SELECT date FROM measurements WHERE tool_id=t.id AND machine_id=? ORDER BY date DESC LIMIT 1) as date,
            (SELECT note FROM measurements WHERE tool_id=t.id AND machine_id=? ORDER BY date DESC LIMIT 1) as note
        FROM tools t
        ORDER BY t.name
    """, (id, id, id)).fetchall()
    
    # Historial
    history = conn.execute("""
        SELECT m.id, m.date, t.name as tool, m.criticality, m.note
        FROM measurements m
        JOIN tools t ON t.id = m.tool_id
        WHERE m.machine_id = ?
        ORDER BY m.date DESC
    """, (id,)).fetchall()
    
    conn.close()
    return render(MACHINE_DETAIL, page_title=machine["name"], machine=machine, current_status=current, history=history)

# ============ HERRAMIENTAS ============

TOOLS_LIST = """
<div class="d-flex justify-content-between align-items-center mb-3">
  <h3>Herramientas de Condición</h3>
  <a class="btn btn-primary" href="/tools/add">+ Agregar</a>
</div>
<p class="text-muted">Define aquí las herramientas para monitorear el estado de las máquinas (ej: Vibración, Temperatura, Ruido, etc.)</p>
<div class="row">
  {% for t in tools %}
  <div class="col-md-4 mb-3">
    <div class="card p-3">
      <h5>{{ t.name }}</h5>
      {% if t.description %}<p class="text-muted small">{{ t.description }}</p>{% endif %}
      <div class="mt-3">
        <a class="btn btn-sm btn-outline-primary" href="/tools/{{ t.id }}/edit">Editar</a>
        <a class="btn btn-sm btn-outline-secondary" href="/tools/{{ t.id }}/status">Ver Estado</a>
        <a class="btn btn-sm btn-outline-danger" href="/tools/{{ t.id }}/delete" onclick="return confirm('¿Eliminar?')">Eliminar</a>
      </div>
    </div>
  </div>
  {% endfor %}
</div>
"""

TOOL_ADD = """
<div class="row justify-content-center">
  <div class="col-md-6">
    <h4>Agregar Herramienta</h4>
    <form method="post" class="card p-4">
      <label class="form-label"><strong>Nombre</strong></label>
      <input class="form-control mb-3" name="name" placeholder="Ej: Vibración, Temperatura, Ruido, Aceite" required>
      
      <label class="form-label"><strong>Descripción (opcional)</strong></label>
      <textarea name="description" class="form-control mb-3" rows="3" placeholder="Qué mide esta herramienta y cómo interpretarla..."></textarea>
      
      <button class="btn btn-primary w-100">Guardar</button>
      <a href="/tools" class="btn btn-secondary w-100 mt-2">Cancelar</a>
    </form>
  </div>
</div>
"""

TOOL_EDIT = """
<div class="row justify-content-center">
  <div class="col-md-6">
    <h4>Editar Herramienta</h4>
    <form method="post" class="card p-4">
      <label class="form-label"><strong>Nombre</strong></label>
      <input class="form-control mb-3" name="name" value="{{ t.name }}" required>
      
      <label class="form-label"><strong>Descripción</strong></label>
      <textarea name="description" class="form-control mb-3" rows="3">{{ t.description or '' }}</textarea>
      
      <button class="btn btn-primary w-100">Guardar</button>
      <a href="/tools" class="btn btn-secondary w-100 mt-2">Cancelar</a>
    </form>
  </div>
</div>
"""

@app.route("/tools")
def tools_list():
    conn = get_db()
    tools = conn.execute("SELECT * FROM tools ORDER BY name").fetchall()
    conn.close()
    return render(TOOLS_LIST, page_title="Herramientas", tools=tools)


@app.route('/calendar', methods=['GET','POST'])
def calendar():
    conn = get_db()
    tools = conn.execute("SELECT * FROM tools ORDER BY name").fetchall()
    machines = conn.execute("SELECT id, name, machine_type FROM machines ORDER BY COALESCE(machine_type,''), name").fetchall()

    if request.method == 'POST':
        date_in = request.form.get('date')
        note = request.form.get('note','').strip()
        tool_ids = request.form.getlist('tool_id')
        machine_ids = request.form.getlist('machine_id')
        # normalize date storage (store as YYYY-MM-DD 00:00)
        date_val = None
        if date_in:
            date_val = f"{date_in} 00:00"

        if not tool_ids or not machine_ids:
          flash('Selecciona al menos una herramienta y una máquina')
          conn.close()
          return redirect('/calendar')

        severity = request.form.get('severity', 'gris')
        # map severity to repair_time
        repair_map = {
          'rojo': '24h',
          'naranja': '48h',
          'amarillo': '72h',
          'verde': 'Sin acción',
          'gris': 'No aplica'
        }
        repair_time = repair_map.get(severity, 'No aplica')

        inserted = 0
        for mid in machine_ids:
          for tid in tool_ids:
            try:
              conn.execute("INSERT INTO measurements (machine_id, tool_id, date, criticality, note, severity, repair_time) VALUES (?,?,?,?,?,?,?)", (int(mid), int(tid), date_val, None, note, severity, repair_time))
              inserted += 1
            except Exception:
              pass
        conn.commit()
        conn.close()
        flash(f'Notas añadidas: {inserted}')
        return redirect('/calendar')

    from datetime import date as _date
    today = _date.today().isoformat()
    # fetch recent notes to display
    conn = get_db()
    recent = conn.execute("""
      SELECT m.date, mac.name as machine, mac.hac_code as hac, t.name as tool, m.severity, m.repair_time, m.note
      FROM measurements m
      LEFT JOIN machines mac ON mac.id = m.machine_id
      LEFT JOIN tools t ON t.id = m.tool_id
      ORDER BY m.date DESC
      LIMIT 100
    """).fetchall()
    # convert rows to simple dict-like
    recent_list = []
    for r in recent:
      recent_list.append({
        'date': r['date'], 'machine': r['machine'], 'hac': r['hac'], 'tool': r['tool'], 'severity': r['severity'], 'repair_time': r['repair_time'], 'note': r['note']
      })
    conn.close()
    return render(CALENDAR_TEMPLATE, page_title='Calendario', tools=tools, machines=machines, today=today, recent=recent_list)

@app.route("/tools/add", methods=["GET","POST"])
def tools_add():
    if request.method == "POST":
        name = request.form.get("name","").strip()
        description = request.form.get("description","").strip()
        conn = get_db()
        try:
            conn.execute("INSERT INTO tools (name, description) VALUES (?,?)", (name, description))
            conn.commit()
            conn.close()
            return redirect("/tools")
        except:
            conn.close()
            flash("Herramienta duplicada o error")
            return render(TOOL_ADD, page_title="Agregar Herramienta")
    return render(TOOL_ADD, page_title="Agregar Herramienta")

@app.route("/tools/<int:id>/edit", methods=["GET","POST"])
def tools_edit(id):
    conn = get_db()
    t = conn.execute("SELECT * FROM tools WHERE id=?", (id,)).fetchone()
    if not t:
        conn.close()
        return "No encontrada", 404
    if request.method == "POST":
        name = request.form.get("name","").strip()
        description = request.form.get("description","").strip()
        conn.execute("UPDATE tools SET name=?, description=? WHERE id=?", (name, description, id))
        conn.commit()
        conn.close()
        return redirect("/tools")
    conn.close()
    return render(TOOL_EDIT, page_title="Editar", t=t)

@app.route("/tools/<int:id>/delete")
def tools_delete(id):
    conn = get_db()
    conn.execute("DELETE FROM measurements WHERE tool_id=?", (id,))
    conn.execute("DELETE FROM tools WHERE id=?", (id,))
    conn.commit()
    conn.close()
    return redirect("/tools")

@app.route("/tools/<int:id>/status")
def tools_status(id):
    conn = get_db()
    tool = conn.execute("SELECT * FROM tools WHERE id=?", (id,)).fetchone()
    if not tool:
        conn.close()
        return "No encontrada", 404
    
    # Obtener última medición de esta herramienta en cada máquina
    machines = conn.execute("""
        SELECT DISTINCT
            m.id as machine_id, m.name,
            (SELECT criticality FROM measurements WHERE tool_id=? AND machine_id=m.id ORDER BY date DESC LIMIT 1) as criticality,
            (SELECT date FROM measurements WHERE tool_id=? AND machine_id=m.id ORDER BY date DESC LIMIT 1) as date,
            (SELECT note FROM measurements WHERE tool_id=? AND machine_id=m.id ORDER BY date DESC LIMIT 1) as note
        FROM machines m
        ORDER BY m.priority DESC, m.name
    """, (id, id, id)).fetchall()
    
    conn.close()
    
    tool_name = tool['name']
    tool_desc = f'<p class="text-muted">{tool["description"]}</p>' if tool['description'] else ''
    
    html = f"""
    <!doctype html>
    <html>
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>Monitor</title>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
        <style>
            body {{ background:#f5f5f5; font-family: system-ui; }}
            .navbar {{ background: linear-gradient(90deg, #2563eb, #1e40af); }}
            .card {{ border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
            .table thead th {{ background: #2563eb; color: white; }}
            .crit-1 {{ background: #d1fae5; color: #047857; }}
            .crit-2 {{ background: #dbeafe; color: #1e40af; }}
            .crit-3 {{ background: #fef3c7; color: #b45309; }}
            .crit-4 {{ background: #fee2e2; color: #991b1b; }}
        </style>
    </head>
    <body>
    <nav class="navbar navbar-expand-lg navbar-dark">
        <div class="container-fluid">
            <a class="navbar-brand fw-bold" href="/">📊 Monitor</a>
            <div class="navbar-nav ms-auto">
                <a class="nav-link" href="/">Máquinas</a>
                <a class="nav-link" href="/tools">Herramientas</a>
            </div>
        </div>
    </nav>
    
    <div class="container mt-4">
        <div class="d-flex justify-content-between align-items-center mb-4">
            <div>
                <h3>{tool_name}</h3>
                {tool_desc}
            </div>
            <a class="btn btn-secondary" href="/tools">Atrás</a>
        </div>
        
        <h5 class="mb-3">Estado actual en todas las máquinas</h5>
        <div class="table-responsive">
            <table class="table">
                <thead>
                    <tr><th>Máquina</th><th>Criticidad</th><th>Fecha</th><th>Nota</th><th></th></tr>
                </thead>
                <tbody>
    """
    
    for m in machines:
        crit = m['criticality']
        if crit:
            crit_class = f"crit-{min(max(crit//3, 1), 4)}"
            html += f"""
                <tr>
                    <td><strong>{m['name']}</strong></td>
                    <td><span class="{crit_class}" style="padding: 5px 10px; border-radius: 5px;">{crit}/10</span></td>
                    <td>{m['date']}</td>
                    <td>{m['note'] or ''}</td>
                    <td><a href="/machines/{m['machine_id']}" class="btn btn-sm btn-outline-primary">Ver</a></td>
                </tr>
            """
        else:
            html += f"""
                <tr class="table-light">
                    <td><strong>{m['name']}</strong></td>
                    <td colspan="3" class="text-muted"><em>Sin mediciones</em></td>
                    <td><a href="/machines/{m['machine_id']}" class="btn btn-sm btn-outline-primary">Ver</a></td>
                </tr>
            """
    
    html += """
                </tbody>
            </table>
        </div>
    </div>
    </body>
    </html>
    """
    return html

# ============ MEDICIONES ============

MEASUREMENT_ADD = """
<div class="row justify-content-center">
  <div class="col-md-6">
    <h4>Registrar Medición - {{ machine.name }}</h4>
    <form method="post" class="card p-4">
      <label class="form-label"><strong>Herramienta</strong></label>
      <select name="tool_id" class="form-select mb-3" required>
        <option value="">-- Seleccionar --</option>
        {% for t in tools %}
        <option value="{{ t.id }}">{{ t.name }}</option>
        {% endfor %}
      </select>
      
      <label class="form-label"><strong>Criticidad (1-10)</strong></label>
      <input type="number" name="criticality" min="1" max="10" class="form-control mb-3" required>
      <small class="text-muted d-block mb-3">1=Bien | 5=Moderado | 10=Peligro</small>
      
      <label class="form-label"><strong>Nota (opcional)</strong></label>
      <textarea name="note" class="form-control mb-3" rows="3"></textarea>
      
      <button class="btn btn-primary w-100">Guardar Medición</button>
    </form>
  </div>
</div>
"""

@app.route("/measurements/add", methods=["GET","POST"])
def measurements_add():
    mid = request.args.get("mid") or request.form.get("machine_id")
    conn = get_db()
    machine = conn.execute("SELECT * FROM machines WHERE id=?", (mid,)).fetchone()
    if not machine:
        conn.close()
        return "Máquina no encontrada", 404
    
    if request.method == "POST":
        tool_id = request.form.get("tool_id")
        criticality = int(request.form.get("criticality"))
        note = request.form.get("note","").strip()
        date = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        conn.execute(
            "INSERT INTO measurements (machine_id, tool_id, date, criticality, note) VALUES (?,?,?,?,?)",
            (mid, tool_id, date, criticality, note)
        )
        conn.commit()
        conn.close()
        return redirect(f"/machines/{mid}")
    
    tools = conn.execute("SELECT * FROM tools ORDER BY name").fetchall()
    conn.close()
    return render(MEASUREMENT_ADD, page_title="Medición", machine=machine, tools=tools)

@app.route("/measurements/<int:id>/delete")
def measurements_delete(id):
    conn = get_db()
    m = conn.execute("SELECT * FROM measurements WHERE id=?", (id,)).fetchone()
    mid = m["machine_id"] if m else 0
    conn.execute("DELETE FROM measurements WHERE id=?", (id,))
    conn.commit()
    conn.close()
    return redirect(f"/machines/{mid}")


# ---- Importar desde Excel (reglas en hoja 'Criterios') ----
def detect_header_row(excel_path, sheet_name, max_scan=10):
    try:
        xls = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
    except Exception:
        return 0
    for i in range(min(max_scan, len(xls))):
        row = xls.iloc[i].astype(str).str.upper().tolist()
        if any('AREA' == v or (isinstance(v, str) and v.strip().startswith('AREA')) for v in row if v and v != 'nan'):
            return i
    for i in range(min(max_scan, len(xls))):
        row = xls.iloc[i].astype(str).str.upper().tolist()
        if any('CÓDIGO' in v or 'CODIGO' in v or 'DENOMIN' in v for v in row if v and v != 'nan'):
            return i
    return 0


def parse_criteria(excel_path):
    try:
        crit = pd.read_excel(excel_path, sheet_name='Criterios', header=None)
    except Exception:
        return {}
    mapping = {}
    for _, row in crit.iterrows():
        texts = []
        factors = []
        for col_idx, val in enumerate(row):
            if pd.isna(val):
                continue
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                factors.append((col_idx, float(val)))
            else:
                s = str(val).strip()
                if s:
                    texts.append((col_idx, s))
        for fcol, fval in factors:
            best = None
            best_dist = 999
            for tcol, tval in texts:
                dist = abs(tcol - fcol)
                if dist < best_dist:
                    best = tval
                    best_dist = dist
            if best:
                mapping[best.upper()] = fval
    return mapping


def score_row_by_criteria(row, criteria_map):
    score = 0.0
    matches = []
    joined = ' '.join([str(x) for x in row.values if not pd.isna(x)])
    joined_u = joined.upper()
    for key, factor in criteria_map.items():
        if key and key in joined_u:
            score += factor
            matches.append((key, factor))
    return score, matches


@app.route('/import_excel', methods=['GET'])
def import_excel():
    # Ruta para importar el Excel según reglas de la hoja 'Criterios'
    excel_path = os.path.join(os.path.dirname(__file__), 'Matriz de condición de equipos principales excel.xlsx')
    if not os.path.exists(excel_path):
        return f"Archivo no encontrado: {excel_path}", 404

    # parse (dataframe) and openpyxl workbook for formatting
    header = detect_header_row(excel_path, 'CM Matrix equipos principales')
    try:
      df = pd.read_excel(excel_path, sheet_name='CM Matrix equipos principales', header=header)
      wb = load_workbook(excel_path, data_only=True)
      ws = wb['CM Matrix equipos principales']
    except Exception as e:
      return f"Error leyendo hoja principal: {e}", 500

    criteria_map = parse_criteria(excel_path)
    max_single = max(criteria_map.values()) if criteria_map else 0

    conn = get_db()
    # ensure a tool exists to tag imports
    tool = conn.execute("SELECT * FROM tools WHERE name=?", ('AutoImport',)).fetchone()
    if not tool:
        conn.execute("INSERT INTO tools (name, description) VALUES (?,?)", ('AutoImport', 'Mediciones importadas desde Excel'))
        conn.commit()
        tool = conn.execute("SELECT * FROM tools WHERE name=?", ('AutoImport',)).fetchone()
    tool_id = tool['id']

    # find likely column names and tool columns to read colors from
    cols_upper = [str(c).upper() for c in df.columns]
    codigo_col = None
    denom_col = None
    comments_col = None
    machine_type_col = None
    tool_names = ['VOSOA','RUTI','COR','ACEITE','VIB','TERMO','DES','DUREZA','EMD','VT / LP','PM','UT','ESPESOR']
    tool_cols_idx = []
    for i, cu in enumerate(cols_upper):
      if 'CÓDIGO' in cu or 'CODIGO' in cu:
        codigo_col = df.columns[i]
      if 'DENOMIN' in cu:
        denom_col = df.columns[i]
      if 'TIPO' in cu or 'TIPO EQU' in cu:
        machine_type_col = df.columns[i]
      if 'COMENT' in cu or 'OBSERV' in cu:
        comments_col = df.columns[i]
      # detect tool columns
      for tn in tool_names:
        if tn in cu:
          tool_cols_idx.append(i)
          break

    inserted = 0
    # compute excel header row (1-based) and starting data row
    excel_header_row = header + 1
    data_start_row = excel_header_row + 1
    for idx, row in df.iterrows():
        name = None
        code = None
        notes = ''
        try:
            if denom_col:
                name = str(row.get(denom_col, '')).strip()
            if codigo_col and pd.notna(row.get(codigo_col)):
                code = str(row.get(codigo_col)).strip()
            if comments_col and pd.notna(row.get(comments_col)):
                notes = str(row.get(comments_col)).strip()
        except Exception:
            continue

        if not name and not code:
            continue

        # determine color from excel cell fills if possible
        detected_color = None
        detected_hex = None
        # corresponding excel row number
        excel_row_num = data_start_row + idx
        # check all columns for cell fill colors (some headers are merged and show as Unnamed)
        for col_idx in range(len(df.columns)):
          try:
            cell = ws.cell(row=excel_row_num, column=col_idx+1)
            fg = None
            if cell.fill and hasattr(cell.fill, 'fgColor'):
              fg = cell.fill.fgColor.rgb or cell.fill.start_color.index
            if fg:
              # normalize hex like 'FF00FF00' or '00FF00'
              hexv = str(fg)
              if len(hexv) == 8 and hexv.startswith('FF'):
                hex6 = hexv[2:]
              elif len(hexv) >= 6:
                hex6 = hexv[-6:]
              else:
                hex6 = None
              if hex6:
                try:
                  r = int(hex6[0:2], 16)
                  g = int(hex6[2:4], 16)
                  b = int(hex6[4:6], 16)
                except Exception:
                  r,g,b = 0,0,0
                # decide color by dominant channel and keep exact hex
                hex_exact = hex6
                # ignore pure white/black
                if (r, g, b) in ((0,0,0),(255,255,255)):
                  pass
                elif r > 200 and g < 120 and b < 120:
                  detected_color = 'red'
                  detected_hex = hex_exact
                  break
                elif r > 200 and g > 150 and b < 150:
                  detected_color = 'yellow'
                  detected_hex = hex_exact
                elif b > max(r,g) and b > 140:
                  detected_color = 'blue'
                  detected_hex = hex_exact
                elif g > max(r,b) and g > 140:
                  detected_color = 'green'
                  detected_hex = hex_exact
          except Exception:
            continue

        # fallback to criteria-based score if no color detected
        score, matches = score_row_by_criteria(row, criteria_map)
        # scale to 0-10
        if max_single > 0:
          denom = max_single * 3
          crit_val = int(round((score / denom) * 10)) if denom > 0 else 0
        else:
          crit_val = 0
        crit_val = max(0, min(10, crit_val))

        # map to priority 1-5
        priority = 1 + (crit_val * 4 // 10)
        # if detected_color, override priority with color mapping and record hex
        if detected_color == 'red':
          priority = 5
        elif detected_color == 'yellow':
          priority = 4
        elif detected_color == 'blue':
          priority = 3
        elif detected_color == 'green':
          priority = 1

        # upsert machine by name or code
        existing = None
        if code:
          existing = conn.execute("SELECT * FROM machines WHERE hac_code=? OR name LIKE ?", (code, f"%{code}%")).fetchone()
        if not existing and name:
            existing = conn.execute("SELECT * FROM machines WHERE name=?", (name,)).fetchone()

        # determine machine_type value from the row if present
        machine_type_val = None
        try:
          if machine_type_col and pd.notna(row.get(machine_type_col)):
            machine_type_val = str(row.get(machine_type_col)).strip()
        except Exception:
          machine_type_val = None

        if existing:
          mid = existing['id']
          # update priority/notes/color unconditionally to reflect Excel exactly
          try:
            conn.execute("UPDATE machines SET priority=?, notes=? WHERE id=?", (priority, (existing['notes'] or '') + ('\n' + notes if notes else ''), mid))
            # update color and color_hex
            if detected_color:
              conn.execute("UPDATE machines SET color=?, color_hex=? WHERE id=?", (detected_color, detected_hex, mid))
            else:
              conn.execute("UPDATE machines SET color=?, color_hex=? WHERE id=?", (None, None, mid))
            # update machine_type if available
            if machine_type_val:
              conn.execute("UPDATE machines SET machine_type=? WHERE id=?", (machine_type_val, mid))
            # update hac_code if present
            if code:
              conn.execute("UPDATE machines SET hac_code=? WHERE id=?", (code, mid))
          except Exception:
            pass
        else:
          # insert
          name_to_insert = name or code
          cur = conn.execute("INSERT INTO machines (name, notes, priority, machine_group, color, color_hex, machine_type, hac_code) VALUES (?,?,?,?,?,?,?,?)", (name_to_insert, notes, priority, 1, detected_color, detected_hex, machine_type_val, code))
          mid = cur.lastrowid

        # insert a measurement marking the computed criticity
        date = datetime.now().strftime("%Y-%m-%d %H:%M")
        matches_note = ';'.join([f"{m[0]}:{m[1]}" for m in matches])
        try:
          conn.execute("INSERT INTO measurements (machine_id, tool_id, date, criticality, note) VALUES (?,?,?,?,?)", (mid, tool_id, date, crit_val, matches_note))
          inserted += 1
        except Exception:
          pass

    conn.commit()
    conn.close()
    return f"Import completado. {inserted} mediciones creadas.", 200

if __name__ == "__main__":
    print("🚀 Monitor de Condición")
    print("➜ http://127.0.0.1:5000")
    app.run(debug=True)
